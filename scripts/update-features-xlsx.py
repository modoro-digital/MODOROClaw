# -*- coding: utf-8 -*-
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = u'T\u00ednh n\u0103ng 9BizClaw'

header_font = Font(bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
cat_fills = {
    'MKT & SALES': PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid'),
    'Assistant': PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid'),
    'Admin': PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid'),
    'Platform': PatternFill(start_color='F3E5F5', end_color='F3E5F5', fill_type='solid'),
    'Roadmap': PatternFill(start_color='ECEFF1', end_color='ECEFF1', fill_type='solid'),
}
thin_border = Border(
    left=Side(style='thin', color='DDDDDD'),
    right=Side(style='thin', color='DDDDDD'),
    top=Side(style='thin', color='DDDDDD'),
    bottom=Side(style='thin', color='DDDDDD')
)
wrap = Alignment(wrap_text=True, vertical='top')

headers = [u'STT', u'Ph\u00e2n lo\u1ea1i', u'K\u00eanh', u'T\u00ednh n\u0103ng', u'M\u00f4 t\u1ea3 chi ti\u1ebft', u'Tr\u1ea1ng th\u00e1i', u'G\u00f3i']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 24
ws.column_dimensions['D'].width = 32
ws.column_dimensions['E'].width = 70
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 10

features = [
    # === MKT & SALES - Zalo CSKH ===
    ('MKT & SALES', u'Zalo \u2013 CSKH', u'Tr\u1ea3 l\u1eddi t\u1ef1 \u0111\u1ed9ng 24/7',
     u'Bot t\u01b0 v\u1ea5n s\u1ea3n ph\u1ea9m, gi\u00e1 c\u1ea3, ch\u00ednh s\u00e1ch, h\u1ed7 tr\u1ee3 mua h\u00e0ng \u2014 kh\u00f4ng c\u1ea7n nh\u00e2n vi\u00ean tr\u1ef1c',
     'OK', 'Free'),

    ('MKT & SALES', u'Zalo \u2013 CSKH', u'Nh\u1eadn di\u1ec7n gi\u1edbi t\u00ednh',
     u'H\u1ecfi "anh hay ch\u1ecb" tr\u01b0\u1edbc khi x\u01b0ng h\u00f4 v\u1edbi 45+ t\u00ean Vi\u1ec7t kh\u00f3 \u0111o\u00e1n, kh\u00f4ng g\u1ecdi nh\u1ea7m',
     'OK', 'Free'),

    ('MKT & SALES', u'Zalo \u2013 CSKH', u'Nh\u1edb t\u1eebng kh\u00e1ch h\u00e0ng',
     u'T\u1ef1 ghi h\u1ed3 s\u01a1 m\u1ed7i kh\u00e1ch: t\u00ean, l\u1ecbch s\u1eed mua, s\u1edf th\u00edch, li\u00ean h\u1ec7. L\u1ea7n sau nh\u1eafn bi\u1ebft ngay',
     'OK', 'Free'),

    ('MKT & SALES', u'Zalo \u2013 CSKH', u'Nh\u1edb t\u1eebng nh\u00f3m Zalo',
     u'Ghi l\u1ecbch s\u1eed nh\u00f3m, th\u00e0nh vi\u00ean ch\u00ednh, ch\u1ee7 \u0111\u1ec1 th\u01b0\u1eddng th\u1ea3o lu\u1eadn',
     'OK', 'Free'),

    ('MKT & SALES', u'Zalo \u2013 CSKH', u'Ch\u00e0o nh\u00f3m m\u1edbi (1 l\u1ea7n)',
     u'T\u1ef1 gi\u1edbi thi\u1ec7u khi \u0111\u01b0\u1ee3c th\u00eam v\u00e0o nh\u00f3m \u2014 ch\u1ec9 ch\u00e0o 1 l\u1ea7n duy nh\u1ea5t, kh\u00f4ng spam',
     'OK', 'Free'),

    ('MKT & SALES', u'Zalo \u2013 CSKH', u'T\u1eeb ch\u1ed1i \u0111\u00fang ph\u1ea1m vi',
     u'Kh\u00f4ng vi\u1ebft b\u00e0i, kh\u00f4ng code, kh\u00f4ng d\u1ecbch \u2014 ch\u1ec9 t\u01b0 v\u1ea5n v\u1ec1 s\u1ea3n ph\u1ea9m c\u00f4ng ty',
     'OK', 'Free'),

    ('MKT & SALES', u'Zalo \u2013 CSKH', u'Kh\u00f4ng b\u1ecba th\u00f4ng tin',
     u'Kh\u00f4ng c\u00f3 trong t\u00e0i li\u1ec7u th\u00ec b\u00e1o kh\u00e1ch + chuy\u1ec3n CEO, kh\u00f4ng t\u1ef1 ch\u1ebf s\u1ed1 li\u1ec7u',
     'OK', 'Free'),

    ('MKT & SALES', u'Zalo \u2013 CSKH', u'Chuy\u1ec3n ti\u1ebfp CEO t\u1ef1 \u0111\u1ed9ng',
     u'Khi\u1ebfu n\u1ea1i, scam, y\u00eau c\u1ea7u nh\u1ea1y c\u1ea3m \u2192 b\u00e1o CEO qua Telegram ngay l\u1eadp t\u1ee9c',
     'OK', 'Free'),

    ('MKT & SALES', u'Zalo \u2013 CSKH', u'Ch\u1eb7n ng\u01b0\u1eddi (blocklist)',
     u'CEO ra l\u1ec7nh ch\u1eb7n \u2192 ng\u01b0\u1eddi trong danh s\u00e1ch \u0111en nh\u1eafn tin bot kh\u00f4ng tr\u1ea3 l\u1eddi',
     'OK', 'Free'),

    ('MKT & SALES', u'Zalo \u2013 CSKH', u'L\u1ecdc spam t\u1ef1 \u0111\u1ed9ng',
     u'Tin qu\u1ea3ng c\u00e1o, m\u1eddi h\u1ee3p t\u00e1c t\u1eeb shop kh\u00e1c \u2192 im l\u1eb7ng ho\u00e0n to\u00e0n, \u0111\u1ec1 xu\u1ea5t block',
     'OK', 'Free'),

    ('MKT & SALES', u'Zalo \u2013 CSKH', u'L\u1ecdc tin h\u1ec7 th\u1ed1ng nh\u00f3m',
     u'Kh\u00f4ng tr\u1ea3 l\u1eddi "X \u0111\u00e3 th\u00eam Y v\u00e0o nh\u00f3m", "X r\u1eddi nh\u00f3m" \u2014 \u0111\u00e2y l\u00e0 th\u00f4ng b\u00e1o, kh\u00f4ng ph\u1ea3i tin',
     'OK', 'Free'),

    ('MKT & SALES', u'Zalo \u2013 CSKH', u'Ch\u1ed1ng bot \u0111\u00e1nh nhau',
     u'Nh\u1eadn di\u1ec7n 6 d\u1ea5u hi\u1ec7u bot t\u1ef1 \u0111\u1ed9ng, im l\u1eb7ng \u0111\u1ec3 tr\u00e1nh v\u00f2ng l\u1eb7p flood',
     'OK', 'Free'),

    ('MKT & SALES', u'Zalo \u2013 CSKH', u'Ch\u1ebf \u0111\u1ed9 ngo\u00e0i gi\u1edd l\u00e0m',
     u'Ngo\u00e0i gi\u1edd: tr\u1ea3 l\u1eddi ng\u1eafn 1 c\u00e2u. Kh\u00e1ch VIP \u0111\u01b0\u1ee3c tag ri\u00eang v\u1eabn \u0111\u01b0\u1ee3c h\u1ed7 tr\u1ee3 24/7',
     'OK', 'Free'),

    ('MKT & SALES', u'Zalo \u2013 CSKH', u'B\u1ea3o m\u1eadt kh\u00e1ch h\u00e0ng',
     u'Kh\u00f4ng ti\u1ebft l\u1ed9 th\u00f4ng tin kh\u00e1ch A cho kh\u00e1ch B, kh\u00f4ng ti\u1ebft l\u1ed9 n\u1ed9i b\u1ed9 c\u00f4ng ty',
     'OK', 'Free'),

    ('MKT & SALES', u'Zalo \u2013 CSKH', u'Ch\u1ed1ng gi\u1ea3 m\u1ea1o danh t\u00ednh',
     u'Kh\u00e1ch t\u1ef1 x\u01b0ng CEO/admin/c\u00f4ng an \u2192 kh\u00f4ng tin, y\u00eau c\u1ea7u x\u00e1c nh\u1eadn qua k\u00eanh n\u1ed9i b\u1ed9',
     'OK', 'Free'),

    ('MKT & SALES', u'Zalo \u2013 CSKH', u'Ch\u1ed1ng hack prompt',
     u'Nh\u1eadn di\u1ec7n jailbreak, "developer mode", base64 \u2192 t\u1eeb ch\u1ed1i, kh\u00f4ng gi\u1ea3i th\u00edch',
     'OK', 'Free'),

    ('MKT & SALES', u'Zalo \u2013 CSKH', u'3 ch\u1ebf \u0111\u1ed9 cho m\u1ed7i nh\u00f3m',
     u'M\u1ed7i nh\u00f3m Zalo ch\u1ecdn ri\u00eang: ch\u1ec9 khi @mention, m\u1ecdi tin nh\u1eafn, ho\u1eb7c t\u1eaft h\u1eb3n',
     'OK', 'Free'),

    ('MKT & SALES', u'Zalo \u2013 CSKH', u'Ch\u00ednh s\u00e1ch ng\u01b0\u1eddi l\u1ea1',
     u'Ng\u01b0\u1eddi ch\u01b0a l\u00e0 b\u1ea1n b\u00e8: tr\u1ea3 l\u1eddi b\u00ecnh th\u01b0\u1eddng / ch\u1ec9 ch\u00e0o 1 l\u1ea7n / im l\u1eb7ng \u2014 CEO t\u00f9y ch\u1ecdn',
     'OK', 'Free'),

    ('MKT & SALES', u'Zalo \u2013 CSKH', u'Ch\u1ed1ng tin tr\u00f9ng l\u1eb7p',
     u'Zalo \u0111\u00f4i khi g\u1eedi 2 event gi\u1ed1ng nhau trong v\u00e0i ms \u2192 bot ch\u1ec9 x\u1eed l\u00fd 1 l\u1ea7n',
     'OK', 'Free'),

    ('MKT & SALES', u'Zalo \u2013 CSKH', u'L\u1ecdc n\u1ed9i dung nh\u1ea1y c\u1ea3m',
     u'T\u1ef1 ch\u1eb7n file path, API key, stack trace, suy ngh\u0129 n\u1ed9i b\u1ed9 tr\u01b0\u1edbc khi g\u1eedi kh\u00e1ch (19 b\u1ed9 l\u1ecdc)',
     'OK', 'Free'),

    ('MKT & SALES', u'Zalo \u2013 CSKH', u'T\u1ef1 chia tin d\u00e0i',
     u'Tin > 800 k\u00fd t\u1ef1 t\u1ef1 \u0111\u1ed9ng chia th\u00e0nh nhi\u1ec1u tin t\u1ea1i d\u1ea5u c\u00e2u ho\u1eb7c \u0111o\u1ea1n',
     'OK', 'Free'),

    # === Assistant - Telegram CEO ===
    ('Assistant', u'Telegram \u2013 CEO', u'Tr\u1ee3 l\u00fd ri\u00eang cho CEO',
     u'Nh\u1eadn l\u1ec7nh b\u1ea5t k\u1ef3: nghi\u00ean c\u1ee9u, so\u1ea1n th\u1ea3o, ph\u00e2n t\u00edch, t\u00f3m t\u1eaft, l\u00ean k\u1ebf ho\u1ea1ch',
     'OK', 'Free'),

    ('Assistant', u'Telegram \u2013 CEO', u'C\u1ea3nh b\u00e1o t\u1ee9c th\u1eddi',
     u'Khi\u1ebfu n\u1ea1i, scam, y\u00eau c\u1ea7u l\u1ea1 t\u1eeb Zalo \u2192 Telegram CEO trong v\u00e0i gi\u00e2y',
     'OK', 'Free'),

    ('Assistant', u'Telegram \u2013 CEO', u'B\u00e1o c\u00e1o t\u1ef1 \u0111\u1ed9ng',
     u'B\u00e1o c\u00e1o s\u00e1ng/t\u1ed1i g\u1eedi qua Telegram theo l\u1ecbch \u0111\u00e3 c\u00e0i',
     'OK', 'Free'),

    ('Assistant', u'Telegram \u2013 CEO', u'\u0110i\u1ec1u khi\u1ec3n bot t\u1eeb xa',
     u'T\u1ea1m d\u1eebng, ti\u1ebfp t\u1ee5c, c\u1ea5u h\u00ecnh qua Dashboard ho\u1eb7c nh\u1eafn l\u1ec7nh Telegram',
     'OK', 'Free'),

    ('Assistant', u'Telegram \u2013 CEO', u'G\u1eedi Zalo t\u1eeb Telegram',
     u'CEO ra l\u1ec7nh g\u1eedi tin Zalo (nh\u00f3m ho\u1eb7c c\u00e1 nh\u00e2n) \u2014 bot h\u1ecfi x\u00e1c nh\u1eadn tr\u01b0\u1edbc khi g\u1eedi',
     u'\u0110ang fix', 'Free'),

    ('Assistant', u'Telegram \u2013 CEO', u'Qu\u1ea3n l\u00fd Zalo t\u1eeb xa',
     u'B\u1eadt/t\u1eaft nh\u00f3m, ch\u1eb7n user, xem tr\u1ea1ng th\u00e1i Zalo t\u1eeb Telegram kh\u00f4ng c\u1ea7n m\u1edf app',
     u'\u0110ang fix', 'Free'),

    ('Assistant', u'Telegram \u2013 CEO', u'Th\u01b0 vi\u1ec7n k\u1ef9 n\u0103ng',
     u'19 prompt m\u1eabu s\u1eb5n: vi\u1ebft content, ph\u00e2n t\u00edch, t\u01b0 v\u1ea5n, so\u1ea1n t\u00e0i li\u1ec7u \u2014 copy-paste d\u00f9ng ngay',
     'OK', 'Free'),

    # === Assistant - Cron ===
    ('Assistant', u'Cron \u2013 T\u1ef1 \u0111\u1ed9ng h\u00f3a', u'B\u00e1o c\u00e1o s\u00e1ng 8:00',
     u'T\u00f3m t\u1eaft ho\u1ea1t \u0111\u1ed9ng h\u00f4m qua, l\u1ecbch h\u00f4m nay, tin nh\u1eafn n\u1ed5i b\u1eadt c\u1ea7n x\u1eed l\u00fd',
     'OK', 'Free'),

    ('Assistant', u'Cron \u2013 T\u1ef1 \u0111\u1ed9ng h\u00f3a', u'T\u00f3m t\u1eaft t\u1ed1i 17:30',
     u'Recap c\u1ea3 ng\u00e0y, vi\u1ec7c c\u00f2n t\u1ed3n \u0111\u1ecdng, kh\u00e1ch c\u1ea7n follow up',
     'OK', 'Free'),

    ('Assistant', u'Cron \u2013 T\u1ef1 \u0111\u1ed9ng h\u00f3a', u'T\u1ef1 t\u1ea1o l\u1ecbch b\u1ea5t k\u1ef3',
     u'CEO t\u1ef1 t\u1ea1o l\u1ecbch t\u1ef1 \u0111\u1ed9ng v\u1edbi prompt t\u00f9y \u00fd t\u1eeb Dashboard',
     'OK', 'Free'),

    ('Assistant', u'Cron \u2013 T\u1ef1 \u0111\u1ed9ng h\u00f3a', u'G\u1eedi c\u1ea3 2 k\u00eanh',
     u'K\u1ebft qu\u1ea3 cron g\u1eedi c\u1ea3 Telegram + Zalo c\u00f9ng l\u00fac',
     'OK', 'Free'),

    ('Assistant', u'Cron \u2013 T\u1ef1 \u0111\u1ed9ng h\u00f3a', u'C\u1ea3nh b\u00e1o l\u1ed7i cron',
     u'Cron th\u1ea5t b\u1ea1i \u2192 CEO nh\u1eadn c\u1ea3nh b\u00e1o ngay, kh\u00f4ng bao gi\u1edd l\u1ed7i \u00e2m th\u1ea7m',
     'OK', 'Free'),

    ('Assistant', u'Cron \u2013 T\u1ef1 \u0111\u1ed9ng h\u00f3a', u'Nh\u1eafc follow-up 9:30',
     u'M\u1ed7i s\u00e1ng t\u1ef1 nh\u1eafc kh\u00e1ch c\u1ea7n follow up d\u1ef1a tr\u00ean l\u1ecbch s\u1eed',
     'OK', 'Free'),

    ('Assistant', u'Cron \u2013 T\u1ef1 \u0111\u1ed9ng h\u00f3a', u'Ki\u1ec3m tra s\u1ee9c kh\u1ecfe 30 ph\u00fat',
     u'M\u1ed7i 30 ph\u00fat ki\u1ec3m tra gateway c\u00f2n s\u1ed1ng, t\u1ef1 restart n\u1ebfu ch\u1ebft',
     'OK', 'Free'),

    # === Admin - Knowledge ===
    ('Admin', u'Knowledge Base', u'Upload t\u00e0i li\u1ec7u',
     u'CEO upload PDF, Word, Excel v\u00e0o 3 th\u01b0 m\u1ee5c: c\u00f4ng ty, s\u1ea3n ph\u1ea9m, nh\u00e2n vi\u00ean',
     'OK', 'Free'),

    ('Admin', u'Knowledge Base', u'Bot \u0111\u1ecdc khi tr\u1ea3 l\u1eddi',
     u'Kh\u00f4ng b\u1ecba th\u00f4ng tin \u2014 ch\u1ec9 tr\u1ea3 l\u1eddi d\u1ef1a tr\u00ean t\u00e0i li\u1ec7u \u0111\u00e3 upload',
     'OK', 'Free'),

    ('Admin', u'Knowledge Base', u'T\u1ef1 index khi kh\u1edfi \u0111\u1ed9ng',
     u'File upload l\u00fac DB l\u1ed7i v\u1eabn \u0111\u01b0\u1ee3c index l\u1ea1i t\u1ef1 \u0111\u1ed9ng khi s\u1eeda xong',
     'OK', 'Free'),

    ('Admin', u'Knowledge Base', u'AI t\u00f3m t\u1eaft n\u1ed9i dung',
     u'Upload file \u2192 AI t\u1ef1 t\u00f3m t\u1eaft v\u00e0o index.md \u0111\u1ec3 bot truy xu\u1ea5t nhanh h\u01a1n',
     'OK', 'Free'),

    # === Admin - B\u1ed9 nh\u1edb ===
    ('Admin', u'B\u1ed9 nh\u1edb', u'H\u1ed3 s\u01a1 kh\u00e1ch Zalo',
     u'T\u00ean, gi\u1edbi t\u00ednh, l\u1ecbch s\u1eed mua, s\u1edf th\u00edch, li\u00ean h\u1ec7 \u2014 bot t\u1ef1 ghi sau m\u1ed7i cu\u1ed9c tr\u00f2 chuy\u1ec7n',
     'OK', 'Member'),

    ('Admin', u'B\u1ed9 nh\u1edb', u'H\u1ed3 s\u01a1 nh\u00f3m Zalo',
     u'T\u00ean nh\u00f3m, th\u00e0nh vi\u00ean ch\u00ednh, ch\u1ee7 \u0111\u1ec1 hay th\u1ea3o lu\u1eadn, quy\u1ebft \u0111\u1ecbnh g\u1ea7n \u0111\u00e2y',
     'OK', 'Member'),

    ('Admin', u'B\u1ed9 nh\u1edb', u'Nh\u1eadt k\u00fd CEO h\u00e0ng ng\u00e0y',
     u'Ghi l\u1ea1i m\u1ecdi l\u1ec7nh + k\u1ebft qu\u1ea3 theo ng\u00e0y, d\u1ec5 t\u00ecm l\u1ea1i',
     'OK', 'Free'),

    ('Admin', u'B\u1ed9 nh\u1edb', u'T\u1ef1 d\u1ecdn file c\u0169',
     u'File kh\u00e1ch qu\u00e1 50KB \u2192 t\u1ef1 x\u00f3a ph\u1ea7n c\u0169 nh\u1ea5t, gi\u1eef trong ng\u01b0\u1ee1ng an to\u00e0n',
     'OK', 'Free'),

    ('Admin', u'B\u1ed9 nh\u1edb', u'CEO ghi ch\u00fa cho kh\u00e1ch',
     u'CEO th\u00eam ghi ch\u00fa v\u00e0o h\u1ed3 s\u01a1 kh\u00e1ch t\u1eeb Dashboard, bot \u0111\u1ecdc khi tr\u1ea3 l\u1eddi l\u1ea7n sau',
     'OK', 'Member'),

    # === Admin - Dashboard ===
    ('Admin', u'Dashboard', u'Tr\u1ea1ng th\u00e1i k\u1ebft n\u1ed1i th\u1ef1c',
     u'Ch\u1ea5m xanh/\u0111\u1ecf Telegram + Zalo, ki\u1ec3m tra th\u1eadt s\u1ef1 m\u1ed7i 45 gi\u00e2y (kh\u00f4ng gi\u1ea3)',
     'OK', 'Free'),

    ('Admin', u'Dashboard', u'T\u1ed5ng quan ho\u1ea1t \u0111\u1ed9ng',
     u'8 s\u1ef1 ki\u1ec7n g\u1ea7n nh\u1ea5t + l\u1eddi ch\u00e0o CEO + l\u1ecbch s\u1eafp t\u1edbi + c\u1ea3nh b\u00e1o c\u1ea7n ch\u00fa \u00fd',
     'OK', 'Free'),

    ('Admin', u'Dashboard', u'Qu\u1ea3n l\u00fd l\u1ecbch t\u1ef1 \u0111\u1ed9ng',
     u'T\u1ea1o, s\u1eeda, x\u00f3a l\u1ecbch cron, test ch\u1ea1y ngay, xem l\u1ecbch s\u1eed',
     'OK', 'Free'),

    ('Admin', u'Dashboard', u'Qu\u1ea3n l\u00fd Zalo chi ti\u1ebft',
     u'Sidebar c\u00e0i \u0111\u1eb7t + tab Nh\u00f3m/B\u1ea1n b\u00e8, b\u1eadt/t\u1eaft t\u1eebng nh\u00f3m, t\u1eebng ng\u01b0\u1eddi',
     'OK', 'Free'),

    ('Admin', u'Dashboard', u'T\u1ea1m d\u1eebng t\u1eebng k\u00eanh',
     u'Pause Telegram ho\u1eb7c Zalo ri\u00eang l\u1ebb \u2014 1 click, h\u1eefu \u00edch khi b\u1ea3o tr\u00ec',
     'OK', 'Free'),

    ('Admin', u'Dashboard', u'Upload t\u00e0i li\u1ec7u k\u00e9o th\u1ea3',
     u'K\u00e9o th\u1ea3 file v\u00e0o 3 danh m\u1ee5c Knowledge, bot \u0111\u1ecdc ngay',
     'OK', 'Free'),

    ('Admin', u'Dashboard', u'G\u1eedi tin test',
     u'B\u1ea5m 1 n\u00fat \u2192 nh\u1eadn tin Telegram \u0111\u1ec3 x\u00e1c nh\u1eadn bot \u0111ang ch\u1ea1y',
     'OK', 'Free'),

    ('Admin', u'Dashboard', u'Xem OpenClaw trong app',
     u'Trang qu\u1ea3n l\u00fd gateway nh\u00fang ngay trong Dashboard, kh\u00f4ng c\u1ea7n m\u1edf tr\u00ecnh duy\u1ec7t',
     'OK', 'Free'),

    ('Admin', u'Dashboard', u'Xem 9Router trong app',
     u'Trang qu\u1ea3n l\u00fd AI provider ngay trong Dashboard, m\u1eadt kh\u1ea9u m\u1eb7c \u0111\u1ecbnh 123456',
     'OK', 'Free'),

    ('Admin', u'Dashboard', u'Prompt m\u1eabu cho CEO',
     u'19 prompt copy-paste s\u1eb5n theo 7 danh m\u1ee5c k\u1ef9 n\u0103ng, d\u00f9ng ngay tr\u00ean Telegram',
     'OK', 'Free'),

    ('Admin', u'Dashboard', u'T\u00f9y ch\u1ec9nh gi\u1ecdng \u0111i\u1ec7u bot',
     u'Ch\u1ecdn m\u1ee9c \u0111\u1ed9 trang tr\u1ecdng, c\u00e1ch x\u01b0ng h\u00f4, t\u00ean bot, ng\u00e0nh ngh\u1ec1',
     'OK', 'Free'),

    ('Admin', u'Dashboard', u'Giao di\u1ec7n s\u00e1ng/t\u1ed1i',
     u'Chuy\u1ec3n \u0111\u1ed5i dark/light theme, thi\u1ebft k\u1ebf premium',
     'OK', 'Free'),

    # === Platform ===
    ('Platform', u'C\u00e0i \u0111\u1eb7t', u'H\u01b0\u1edbng d\u1eabn 4 b\u01b0\u1edbc',
     u'Telegram token \u2192 AI provider \u2192 th\u00f4ng tin c\u00f4ng ty \u2192 Zalo QR. Xong l\u00e0 ch\u1ea1y',
     'OK', 'Free'),

    ('Platform', u'C\u00e0i \u0111\u1eb7t', u'Kh\u00f4ng c\u1ea7n c\u00e0i g\u00ec th\u00eam',
     u'EXE (Windows) + DMG (Mac) \u0111\u00e3 g\u1ed3m s\u1eb5n Node.js v\u00e0 4 package c\u1ea7n thi\u1ebft',
     'OK', 'Free'),

    ('Platform', u'C\u00e0i \u0111\u1eb7t', u'Thanh ti\u1ebfn \u0111\u1ed9 l\u1ea7n \u0111\u1ea7u',
     u'L\u1ea7n \u0111\u1ea7u m\u1edf app gi\u1ea3i n\u00e9n 140K file \u2192 hi\u1ec7n thanh ti\u1ebfn \u0111\u1ed9 r\u00f5 r\u00e0ng (Windows)',
     'OK', 'Free'),

    ('Platform', u'C\u00e0i \u0111\u1eb7t', u'T\u1ef1 v\u00e1 plugin m\u1ed7i l\u1ea7n m\u1edf',
     u'M\u1ed7i kh\u1edfi \u0111\u1ed9ng t\u1ef1 \u0111\u1ed9ng v\u00e1 openzalo (blocklist, pause, dedup, filter...)',
     'OK', 'Free'),

    ('Platform', u'C\u00e0i \u0111\u1eb7t', u'Build t\u1ef1 \u0111\u1ed9ng Mac + Win',
     u'GitHub Actions build DMG (arm64 + x64) v\u00e0 EXE t\u1ef1 \u0111\u1ed9ng khi c\u00f3 b\u1ea3n m\u1edbi',
     'OK', 'Free'),

    ('Platform', u'C\u00e0i \u0111\u1eb7t', u'T\u1ef1 s\u1eeda c\u1ea5u h\u00ecnh l\u1ed7i',
     u'C\u1ea5u h\u00ecnh b\u1ecb l\u1ed7i schema \u2192 t\u1ef1 ph\u00e1t hi\u1ec7n + x\u00f3a key sai + th\u1eed l\u1ea1i t\u1ef1 \u0111\u1ed9ng',
     'OK', 'Free'),

    ('Platform', u'C\u00e0i \u0111\u1eb7t', u'Kh\u00f4ng spam tin kh\u1edfi \u0111\u1ed9ng',
     u'Th\u00f4ng b\u00e1o "s\u1eb5n s\u00e0ng" ch\u1ec9 g\u1eedi 1 l\u1ea7n m\u1ed7i 30 ph\u00fat, nh\u1edb qua l\u1ea7n restart',
     'OK', 'Free'),

    ('Platform', u'C\u00e0i \u0111\u1eb7t', u'Ch\u1ed1ng Mac ng\u1ee7 \u0111\u00f4ng',
     u'Ng\u0103n macOS t\u1eaft timer khi app ch\u1ea1y n\u1ec1n \u2014 cron v\u1eabn ch\u1ea1y \u0111\u00fang gi\u1edd',
     'OK', 'Free'),

    ('Platform', u'C\u00e0i \u0111\u1eb7t', u'T\u1ef1 c\u1eadp nh\u1eadt workspace',
     u'B\u1ea3n m\u1edbi t\u1ef1 c\u1eadp nh\u1eadt lu\u1eadt bot + c\u00f4ng c\u1ee5 khi kh\u1edfi \u0111\u1ed9ng, kh\u00f4ng c\u1ea7n c\u00e0i l\u1ea1i',
     'OK', 'Free'),

    # === Roadmap ===
    ('Roadmap', u'(S\u1eafp c\u00f3)', u'Li\u00ean k\u00eanh Facebook',
     u'\u0110\u0103ng b\u00e0i t\u1ef1 \u0111\u1ed9ng l\u00ean Facebook t\u1eeb l\u1ec7nh Telegram',
     u'S\u1eafp ra', '-'),

    ('Roadmap', u'(S\u1eafp c\u00f3)', u'Th\u01b0 m\u1ee5c h\u00ecnh \u1ea3nh',
     u'Qu\u1ea3n l\u00fd h\u00ecnh \u1ea3nh \u0111\u1ec3 li\u00ean k\u1ebft \u0111\u0103ng b\u00e0i marketing',
     u'S\u1eafp ra', '-'),

    ('Roadmap', u'(S\u1eafp c\u00f3)', u'AI t\u1ef1 l\u00e0m h\u00ecnh',
     u'AI t\u1ea1o h\u00ecnh \u1ea3nh cho marketing t\u1ef1 \u0111\u1ed9ng',
     u'S\u1eafp ra', '-'),

    ('Roadmap', u'(S\u1eafp c\u00f3)', u'AI t\u1ef1 l\u00e0m clip',
     u'AI t\u1ea1o video ng\u1eafn cho marketing',
     u'S\u1eafp ra', '-'),

    ('Roadmap', u'(S\u1eafp c\u00f3)', u'C\u1eadp nh\u1eadt 1 click',
     u'Kh\u00e1ch b\u1ea5m 1 n\u00fat \u2192 t\u1ea3i v\u00e0 c\u00e0i b\u1ea3n m\u1edbi nh\u1ea5t t\u1ef1 \u0111\u1ed9ng',
     u'S\u1eafp ra', '-'),

    ('Roadmap', u'(S\u1eafp c\u00f3)', u'\u0110\u1ed3ng b\u1ed9 Google Calendar',
     u'L\u1ecbch h\u1eb9n t\u1ef1 \u0111\u1ed3ng v\u1edbi Google Calendar c\u1ee7a CEO',
     u'S\u1eafp ra', '-'),
]

for i, (cat, channel, name, desc, status, version) in enumerate(features, 1):
    row = i + 1
    ws.cell(row=row, column=1, value=i)
    ws.cell(row=row, column=2, value=cat)
    ws.cell(row=row, column=3, value=channel)
    ws.cell(row=row, column=4, value=name)
    ws.cell(row=row, column=5, value=desc)
    ws.cell(row=row, column=6, value=status)
    ws.cell(row=row, column=7, value=version)

    fill = cat_fills.get(cat, None)
    for col in range(1, 8):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = wrap
        if fill:
            cell.fill = fill
        if col == 6:
            if status == 'OK':
                cell.font = Font(color='16A34A', bold=True)
            elif status.startswith(u'\u0110ang'):
                cell.font = Font(color='D97706', bold=True)
            elif status.startswith(u'S\u1eafp'):
                cell.font = Font(color='6366F1', bold=True)

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:G{len(features)+1}'

wb.save('C:/Users/buitu/Desktop/claw/9BizClaw-tinh-nang.xlsx')
print(f'Done: {len(features)} features')
